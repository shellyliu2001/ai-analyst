import pandas as pd
import os
import glob
import PyPDF2
from openpyxl import load_workbook

# Global variable for data folder path
DATA_FOLDER = "data/input"

def extract_excel_model():
    """Extract model.xlsx from data/input folder."""
    data_folder = DATA_FOLDER
    
    # Read Excel file with openpyxl to get values and formulas
    excel_path = os.path.join(data_folder, "model.xlsx")
    excel_data = None
    excel_formulas = None
    if os.path.exists(excel_path):
        # Read with pandas for DataFrame
        excel_data = pd.read_excel(excel_path)
        
        # Read with openpyxl for values and formulas
        wb = load_workbook(excel_path, data_only=False)
        ws = wb.active
        excel_formulas = {}
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_ref = cell.coordinate
                    excel_formulas[cell_ref] = {
                        'value': cell.value,
                        'formula': cell.value if str(cell.value).startswith('=') else None
                    }
    
    return excel_data, excel_formulas

def extract_pdf_files():
    """Extract all PDF files in data/input folder, extracting text content."""
    data_folder = DATA_FOLDER
    pdf_files = glob.glob(os.path.join(data_folder, "*.pdf"))
    pdf_data = {}
    
    for pdf_file in pdf_files:
        try:
            with open(pdf_file, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text = ""
                
                for page in reader.pages:
                    # Extract text from page
                    page_text = page.extract_text()
                    text += page_text + "\n"
                
                pdf_data[os.path.basename(pdf_file)] = text
                
        except Exception as e:
            print(f"Error reading {pdf_file}: {e}")
            continue
    
    return pdf_data

def extract_tables_from_pdf():
    """
    TODO: Implement table extraction functionality
    This function should be implemented to extract tables from PDF files.
    Consider using libraries like:
    - pdfplumber for better table extraction
    - tabula-py for table-specific extraction
    - camelot-py for advanced table detection